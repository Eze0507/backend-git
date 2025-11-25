from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Sum, Count
from django.shortcuts import get_object_or_404
from decimal import Decimal

from personal_admin.model_nomina import Nomina, DetalleNomina
from personal_admin.models import Empleado, Bitacora
from personal_admin.serializers.serializers_nomina import (
    NominaReadSerializer,
    NominaListSerializer,
    NominaWriteSerializer,
    NominaUpdateEstadoSerializer,
    DetalleNominaReadSerializer,
    DetalleNominaWriteSerializer,
)


def registrar_bitacora(usuario, accion, modulo, descripcion, request=None):
    """
    Función helper para registrar acciones en la bitácora.
    """
    ip_address = None
    if request:
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip_address = x_forwarded_for.split(',')[0]
        else:
            ip_address = request.META.get('REMOTE_ADDR')
    
    user_tenant = None
    if usuario and usuario.is_authenticated and hasattr(usuario, 'profile'):
        user_tenant = usuario.profile.tenant
    
    Bitacora.objects.create(
        usuario=usuario,
        accion=accion,
        modulo=modulo,
        descripcion=descripcion,
        ip_address=ip_address,
        tenant=user_tenant
    )


class NominaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar nóminas.
    
    Endpoints:
    - list: Listar todas las nóminas
    - retrieve: Obtener detalle de una nómina específica
    - create: Crear nueva nómina (genera detalles automáticamente)
    - update/partial_update: Actualizar nómina
    - destroy: Eliminar nómina
    - cambiar_estado: Cambiar el estado de la nómina
    - recalcular: Recalcular todos los detalles de la nómina
    - estadisticas: Obtener estadísticas de nóminas
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['mes', 'estado']
    ordering_fields = ['fecha_inicio', 'fecha_corte', 'mes', 'total_nomina']
    ordering = ['-fecha_inicio']
    
    def get_queryset(self):
        """Filtrar nóminas por tenant del usuario"""
        user_tenant = self.request.user.profile.tenant
        queryset = Nomina.objects.filter(tenant=user_tenant)
        
        # Filtros adicionales
        mes = self.request.query_params.get('mes', None)
        estado = self.request.query_params.get('estado', None)
        año = self.request.query_params.get('año', None)
        
        if mes:
            queryset = queryset.filter(mes=mes)
        if estado:
            queryset = queryset.filter(estado=estado)
        if año:
            queryset = queryset.filter(fecha_inicio__year=año)
        
        return queryset
    
    def get_serializer_class(self):
        """Retornar el serializer apropiado según la acción"""
        if self.action == 'list':
            return NominaListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return NominaWriteSerializer
        elif self.action == 'cambiar_estado':
            return NominaUpdateEstadoSerializer
        return NominaReadSerializer
    
    def get_serializer_context(self):
        """Agregar el request al contexto del serializer"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def perform_create(self, serializer):
        """Crear nómina con tenant y registrar en bitácora"""
        user_tenant = self.request.user.profile.tenant
        instance = serializer.save(tenant=user_tenant)
        
        # Registrar en bitácora
        descripcion = (
            f"Nómina creada para {instance.get_periodo()} "
            f"con {instance.detalles.count()} empleados. "
            f"Total: Bs. {instance.total_nomina}"
        )
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.CREAR,
            modulo=Bitacora.Modulo.EMPLEADO,  # Usar módulo apropiado
            descripcion=descripcion,
            request=self.request
        )
    
    def perform_update(self, serializer):
        """Actualizar nómina y registrar en bitácora"""
        instance = self.get_object()
        estado_original = instance.estado
        
        instance = serializer.save()
        
        # Registrar en bitácora
        cambios = []
        if estado_original != instance.estado:
            cambios.append(f"Estado: {estado_original} → {instance.estado}")
        
        descripcion = f"Nómina {instance.get_periodo()} actualizada"
        if cambios:
            descripcion += f". Cambios: {', '.join(cambios)}"
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.EDITAR,
            modulo=Bitacora.Modulo.EMPLEADO,
            descripcion=descripcion,
            request=self.request
        )
    
    def perform_destroy(self, instance):
        """Eliminar nómina y registrar en bitácora"""
        periodo = instance.get_periodo()
        cantidad_empleados = instance.detalles.count()
        total = instance.total_nomina
        
        instance.delete()
        
        # Registrar en bitácora
        descripcion = (
            f"Nómina {periodo} eliminada. "
            f"Tenía {cantidad_empleados} empleados y total de Bs. {total}"
        )
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.ELIMINAR,
            modulo=Bitacora.Modulo.EMPLEADO,
            descripcion=descripcion,
            request=self.request
        )
    
    @action(detail=True, methods=['patch'])
    def cambiar_estado(self, request, pk=None):
        """
        Endpoint personalizado para cambiar el estado de la nómina.
        POST /api/nominas/{id}/cambiar_estado/
        Body: {"estado": "Pagada"}
        """
        nomina = self.get_object()
        serializer = NominaUpdateEstadoSerializer(
            nomina, 
            data=request.data, 
            partial=True
        )
        
        if serializer.is_valid():
            estado_anterior = nomina.estado
            serializer.save()
            
            # Registrar en bitácora
            descripcion = (
                f"Estado de nómina {nomina.get_periodo()} cambiado: "
                f"{estado_anterior} → {nomina.estado}"
            )
            
            registrar_bitacora(
                usuario=request.user,
                accion=Bitacora.Accion.EDITAR,
                modulo=Bitacora.Modulo.EMPLEADO,
                descripcion=descripcion,
                request=request
            )
            
            return Response(
                NominaReadSerializer(nomina).data,
                status=status.HTTP_200_OK
            )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def recalcular(self, request, pk=None):
        """
        Endpoint para recalcular todos los detalles de la nómina.
        POST /api/nominas/{id}/recalcular/
        """
        nomina = self.get_object()
        
        # Recalcular cada detalle
        detalles_actualizados = 0
        for detalle in nomina.detalles.all():
            detalle.calcular_todos_los_campos()
            detalle.save()
            detalles_actualizados += 1
        
        # Actualizar total de la nómina
        nomina.calcular_total_nomina()
        nomina.save(update_fields=['total_nomina'])
        
        # Registrar en bitácora
        descripcion = (
            f"Nómina {nomina.get_periodo()} recalculada. "
            f"{detalles_actualizados} detalles actualizados. "
            f"Nuevo total: Bs. {nomina.total_nomina}"
        )
        
        registrar_bitacora(
            usuario=request.user,
            accion=Bitacora.Accion.EDITAR,
            modulo=Bitacora.Modulo.EMPLEADO,
            descripcion=descripcion,
            request=request
        )
        
        return Response(
            {
                "message": f"Nómina recalculada exitosamente. {detalles_actualizados} detalles actualizados.",
                "total_nomina": nomina.total_nomina,
                "data": NominaReadSerializer(nomina).data
            },
            status=status.HTTP_200_OK
        )
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """
        Endpoint para obtener estadísticas de nóminas.
        GET /api/nominas/estadisticas/
        """
        user_tenant = request.user.profile.tenant
        
        # Estadísticas generales
        total_nominas = Nomina.objects.filter(tenant=user_tenant).count()
        nominas_pendientes = Nomina.objects.filter(
            tenant=user_tenant, 
            estado=Nomina.Estado.PENDIENTE
        ).count()
        nominas_pagadas = Nomina.objects.filter(
            tenant=user_tenant, 
            estado=Nomina.Estado.PAGADA
        ).count()
        
        # Total pagado en nóminas
        total_pagado = Nomina.objects.filter(
            tenant=user_tenant,
            estado=Nomina.Estado.PAGADA
        ).aggregate(total=Sum('total_nomina'))['total'] or Decimal('0.00')
        
        # Total pendiente
        total_pendiente = Nomina.objects.filter(
            tenant=user_tenant,
            estado=Nomina.Estado.PENDIENTE
        ).aggregate(total=Sum('total_nomina'))['total'] or Decimal('0.00')
        
        # Nómina más reciente
        nomina_reciente = Nomina.objects.filter(
            tenant=user_tenant
        ).order_by('-fecha_inicio').first()
        
        estadisticas = {
            "total_nominas": total_nominas,
            "nominas_pendientes": nominas_pendientes,
            "nominas_pagadas": nominas_pagadas,
            "total_pagado": float(total_pagado),
            "total_pendiente": float(total_pendiente),
            "nomina_reciente": NominaListSerializer(nomina_reciente).data if nomina_reciente else None
        }
        
        return Response(estadisticas, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def exportar_excel(self, request, pk=None):
        """
        Exporta una nómina específica a Excel con todos los detalles.
        GET /api/nominas/{id}/exportar_excel/
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        from datetime import datetime
        
        nomina = self.get_object()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Nómina {nomina.get_periodo()}"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = f"NÓMINA - {nomina.get_periodo().upper()}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Info general
        ws['A3'] = "Taller:"
        ws['B3'] = request.user.profile.tenant.nombre_taller
        ws['A4'] = "Período:"
        ws['B4'] = f"{nomina.fecha_inicio} al {nomina.fecha_corte}"
        ws['A5'] = "Estado:"
        ws['B5'] = nomina.get_estado_display()
        ws['A6'] = "Fecha de registro:"
        ws['B6'] = nomina.fecha_registro.strftime("%Y-%m-%d %H:%M")
        
        # Encabezados de tabla
        headers = ['EMPLEADO', 'CI', 'SUELDO BASE', 'HORAS EXTRAS', 'TOTAL BRUTO', 'DESCUENTOS', 'SUELDO NETO']
        row = 8
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos de empleados
        detalles = nomina.detalles.select_related('empleado').order_by('empleado__apellido')
        row = 9
        for detalle in detalles:
            ws.cell(row=row, column=1, value=f"{detalle.empleado.nombre} {detalle.empleado.apellido}")
            ws.cell(row=row, column=2, value=detalle.empleado.ci)
            ws.cell(row=row, column=3, value=float(detalle.sueldo))
            ws.cell(row=row, column=4, value=float(detalle.horas_extras))
            ws.cell(row=row, column=5, value=float(detalle.total_bruto))
            ws.cell(row=row, column=6, value=float(detalle.total_descuento))
            ws.cell(row=row, column=7, value=float(detalle.sueldo_neto))
            
            # Aplicar bordes
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border
                if col >= 3:  # Formato de moneda para columnas numéricas
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
            
            row += 1
        
        # Totales
        row += 1
        ws.cell(row=row, column=1, value="TOTAL NÓMINA:").font = Font(bold=True)
        ws.cell(row=row, column=7, value=float(nomina.total_nomina)).font = Font(bold=True)
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Nomina_{nomina.get_periodo().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        # Registrar en bitácora
        registrar_bitacora(
            usuario=request.user,
            accion=Bitacora.Accion.CONSULTAR,
            modulo=Bitacora.Modulo.EMPLEADO,
            descripcion=f"Nómina {nomina.get_periodo()} exportada a Excel",
            request=request
        )
        
        return response


class DetalleNominaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar detalles de nómina.
    
    Endpoints:
    - list: Listar todos los detalles (con filtros)
    - retrieve: Obtener detalle específico
    - create: Crear nuevo detalle
    - update/partial_update: Actualizar detalle
    - destroy: Eliminar detalle
    - recalcular: Recalcular un detalle específico
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['empleado__nombre', 'empleado__apellido', 'empleado__ci']
    ordering_fields = ['sueldo_neto', 'total_bruto', 'empleado__apellido']
    ordering = ['empleado__apellido']
    
    def get_queryset(self):
        """Filtrar detalles por tenant del usuario"""
        user_tenant = self.request.user.profile.tenant
        queryset = DetalleNomina.objects.filter(tenant=user_tenant)
        
        # Filtrar por nómina si se proporciona
        nomina_id = self.request.query_params.get('nomina', None)
        if nomina_id:
            queryset = queryset.filter(nomina_id=nomina_id)
        
        # Filtrar por empleado si se proporciona
        empleado_id = self.request.query_params.get('empleado', None)
        if empleado_id:
            queryset = queryset.filter(empleado_id=empleado_id)
        
        return queryset.select_related('nomina', 'empleado', 'empleado__cargo')
    
    def get_serializer_class(self):
        """Retornar el serializer apropiado según la acción"""
        if self.action in ['create', 'update', 'partial_update']:
            return DetalleNominaWriteSerializer
        return DetalleNominaReadSerializer
    
    def perform_create(self, serializer):
        """Crear detalle con tenant y registrar en bitácora"""
        user_tenant = self.request.user.profile.tenant
        instance = serializer.save(tenant=user_tenant)
        
        # Registrar en bitácora
        descripcion = (
            f"Detalle de nómina agregado: {instance.empleado.nombre} {instance.empleado.apellido} "
            f"a nómina {instance.nomina.get_periodo()}. "
            f"Sueldo neto: Bs. {instance.sueldo_neto}"
        )
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.CREAR,
            modulo=Bitacora.Modulo.EMPLEADO,
            descripcion=descripcion,
            request=self.request
        )
    
    def perform_update(self, serializer):
        """Actualizar detalle y registrar en bitácora"""
        instance = self.get_object()
        sueldo_anterior = instance.sueldo_neto
        
        instance = serializer.save()
        
        # Registrar en bitácora
        descripcion = (
            f"Detalle de nómina actualizado: {instance.empleado.nombre} {instance.empleado.apellido} "
            f"en nómina {instance.nomina.get_periodo()}. "
            f"Sueldo neto: Bs. {sueldo_anterior} → Bs. {instance.sueldo_neto}"
        )
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.EDITAR,
            modulo=Bitacora.Modulo.EMPLEADO,
            descripcion=descripcion,
            request=self.request
        )
    
    def perform_destroy(self, instance):
        """Eliminar detalle y registrar en bitácora"""
        empleado_nombre = f"{instance.empleado.nombre} {instance.empleado.apellido}"
        periodo = instance.nomina.get_periodo()
        sueldo = instance.sueldo_neto
        
        instance.delete()
        
        # Registrar en bitácora
        descripcion = (
            f"Detalle de nómina eliminado: {empleado_nombre} "
            f"de nómina {periodo}. Sueldo era: Bs. {sueldo}"
        )
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.ELIMINAR,
            modulo=Bitacora.Modulo.EMPLEADO,
            descripcion=descripcion,
            request=self.request
        )
    
    @action(detail=True, methods=['post'])
    def recalcular(self, request, pk=None):
        """
        Endpoint para recalcular un detalle específico.
        POST /api/detalle-nomina/{id}/recalcular/
        """
        detalle = self.get_object()
        
        # Guardar valores anteriores
        bruto_anterior = detalle.total_bruto
        neto_anterior = detalle.sueldo_neto
        
        # Recalcular
        detalle.calcular_todos_los_campos()
        detalle.save()
        
        # Actualizar total de la nómina padre
        detalle.nomina.calcular_total_nomina()
        detalle.nomina.save(update_fields=['total_nomina'])
        
        # Registrar en bitácora
        descripcion = (
            f"Detalle recalculado: {detalle.empleado.nombre} {detalle.empleado.apellido}. "
            f"Bruto: Bs. {bruto_anterior} → Bs. {detalle.total_bruto}, "
            f"Neto: Bs. {neto_anterior} → Bs. {detalle.sueldo_neto}"
        )
        
        registrar_bitacora(
            usuario=request.user,
            accion=Bitacora.Accion.EDITAR,
            modulo=Bitacora.Modulo.EMPLEADO,
            descripcion=descripcion,
            request=request
        )
        
        return Response(
            {
                "message": "Detalle recalculado exitosamente.",
                "data": DetalleNominaReadSerializer(detalle).data
            },
            status=status.HTTP_200_OK
        )
