# finanzas_facturacion/views_export.py
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.shortcuts import get_object_or_404
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import logging

from .models import Pago

logger = logging.getLogger(__name__)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def exportar_pago_pdf(request, pk):
    """
    Exporta el detalle de un pago en formato PDF
    GET /api/finanzas-facturacion/pagos/{id}/export/pdf/
    """
    try:
        user = request.user
        user_tenant = user.profile.tenant
        
        # Obtener el pago
        pago = get_object_or_404(Pago, id=pk, tenant=user_tenant)
        
        # Crear el buffer para el PDF
        buffer = BytesIO()
        
        # Crear el documento PDF
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#673AB7'),
            spaceAfter=30,
            alignment=1  # Centrado
        )
        
        # Título
        title = Paragraph("Detalle del Pago", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Datos del pago
        data = [
            ['Campo', 'Valor'],
            ['ID', str(pago.id)],
            ['Orden', f"#{pago.orden_trabajo.id}" if pago.orden_trabajo else 'N/A'],
            ['Cliente', pago.orden_trabajo.cliente.nombre if pago.orden_trabajo and pago.orden_trabajo.cliente else 'N/A'],
            ['Monto', f"Bs. {pago.monto:.2f}"],
            ['Método de Pago', pago.metodo_pago],
            ['Estado', pago.estado],
            ['Fecha', pago.fecha_pago.strftime('%Y-%m-%d') if pago.fecha_pago else 'N/A'],
            ['Payment Intent', pago.stripe_payment_intent_id or '—'],
        ]
        
        # Crear tabla
        table = Table(data, colWidths=[2*inch, 4*inch])
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#673AB7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3E5F5')]),
        ]))
        
        elements.append(table)
        
        # Construir el PDF
        doc.build(elements)
        
        # Obtener el valor del buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Crear la respuesta HTTP
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="pago_{pago.id}.pdf"'
        
        logger.info(f"📄 PDF generado para pago #{pago.id}")
        return response
        
    except Exception as e:
        logger.error(f"❌ Error generando PDF: {e}")
        return HttpResponse(f"Error: {str(e)}", status=500)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def exportar_pago_excel(request, pk):
    """
    Exporta el detalle de un pago en formato Excel
    GET /api/finanzas-facturacion/pagos/{id}/export/excel/
    """
    try:
        user = request.user
        user_tenant = user.profile.tenant
        
        # Obtener el pago
        pago = get_object_or_404(Pago, id=pk, tenant=user_tenant)
        
        # Crear el workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Pago {pago.id}"
        
        # Estilos
        header_fill = PatternFill(start_color="673AB7", end_color="673AB7", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezado
        ws['A1'] = 'Campo'
        ws['B1'] = 'Valor'
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        ws['A1'].alignment = header_alignment
        ws['B1'].alignment = header_alignment
        
        # Datos
        data = [
            ['ID', str(pago.id)],
            ['Orden', f"#{pago.orden_trabajo.id}" if pago.orden_trabajo else 'N/A'],
            ['Cliente', pago.orden_trabajo.cliente.nombre if pago.orden_trabajo and pago.orden_trabajo.cliente else 'N/A'],
            ['Monto', f"Bs. {pago.monto:.2f}"],
            ['Método de Pago', pago.metodo_pago],
            ['Estado', pago.estado],
            ['Fecha', pago.fecha_pago.strftime('%Y-%m-%d') if pago.fecha_pago else 'N/A'],
            ['Payment Intent', pago.stripe_payment_intent_id or '—'],
        ]
        
        # Agregar datos a la hoja
        for idx, row in enumerate(data, start=2):
            ws[f'A{idx}'] = row[0]
            ws[f'B{idx}'] = row[1]
            ws[f'A{idx}'].font = Font(bold=True)
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Crear la respuesta HTTP
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="pago_{pago.id}.xlsx"'
        
        logger.info(f"📊 Excel generado para pago #{pago.id}")
        return response
        
    except Exception as e:
        logger.error(f"❌ Error generando Excel: {e}")
        return HttpResponse(f"Error: {str(e)}", status=500)

