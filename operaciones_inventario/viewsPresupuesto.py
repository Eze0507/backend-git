from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q
from django.http import HttpResponse
from .modelsPresupuesto import presupuesto, detallePresupuesto
from .serializers.serializersPresupuesto import PresupuestoSerializer, DetallePresupuestoSerializer
from personal_admin.views import registrar_bitacora
from personal_admin.models import Bitacora

# Imports para reportes
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Preformatted
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class DetallePresupuestoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para CRUD de Detalles de Presupuesto.
    Permite crear, leer, actualizar y eliminar detalles de presupuesto de forma independiente.
    """
    queryset = detallePresupuesto.objects.all()
    serializer_class = DetallePresupuestoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        presupuesto_id = self.request.query_params.get('presupuesto_id', None)
        if presupuesto_id:
            queryset = queryset.filter(presupuesto_id=presupuesto_id)

        item_id = self.request.query_params.get('item_id', None)
        if item_id:
            queryset = queryset.filter(item_id=item_id)

        return queryset.order_by('id')

    def create(self, request, *args, **kwargs):
        """Crear un nuevo detalle de presupuesto"""
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            detalle = serializer.save()
            # Recalcular totales del presupuesto padre
            if detalle.presupuesto:
                detalle.presupuesto.recalcular_totales()
            return Response(DetallePresupuestoSerializer(detalle).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        """Actualizar un detalle de presupuesto existente"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            detalle = serializer.save()
            # Recalcular totales del presupuesto padre
            if detalle.presupuesto:
                detalle.presupuesto.recalcular_totales()
            return Response(DetallePresupuestoSerializer(detalle).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        """Eliminar un detalle de presupuesto"""
        instance = self.get_object()
        presupuesto_padre = instance.presupuesto
        instance.delete()
        # Recalcular totales del presupuesto padre después de eliminar
        if presupuesto_padre:
            presupuesto_padre.recalcular_totales()
        return Response({'message': 'Detalle de presupuesto eliminado correctamente.'}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Crear múltiples detalles de presupuesto a la vez"""
        detalles_data = request.data if isinstance(request.data, list) else [request.data]
        serializer = DetallePresupuestoSerializer(data=detalles_data, many=True)
        if serializer.is_valid():
            detalles = serializer.save()
            # Recalcular totales para todos los presupuestos afectados
            presupuestos_afectados = set()
            for detalle in detalles:
                if detalle.presupuesto:
                    presupuestos_afectados.add(detalle.presupuesto)
            for presup in presupuestos_afectados:
                presup.recalcular_totales()
            return Response(DetallePresupuestoSerializer(detalles, many=True).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PresupuestoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para CRUD de Presupuestos.
    """
    queryset = presupuesto.objects.all().prefetch_related('detalles')
    serializer_class = PresupuestoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        vehiculo_id = self.request.query_params.get('vehiculo_id', None)
        if vehiculo_id:
            queryset = queryset.filter(vehiculo_id=vehiculo_id)

        fecha_inicio = self.request.query_params.get('fecha_inicio', None)
        if fecha_inicio:
            queryset = queryset.filter(fecha_inicio=fecha_inicio)

        return queryset.order_by('-id')

    def perform_create(self, serializer):
        """Crear presupuesto y registrar en bitácora"""
        # Ejecutar la creación original
        instance = serializer.save()
        
        # Preparar información para la bitácora
        cliente_info = f"{instance.cliente.nombre} {instance.cliente.apellido}" if instance.cliente else "Sin cliente"
        vehiculo_info = f"{instance.vehiculo.marca.nombre if instance.vehiculo and instance.vehiculo.marca else ''} {instance.vehiculo.modelo.nombre if instance.vehiculo and instance.vehiculo.modelo else ''} (Placa: {instance.vehiculo.numero_placa if instance.vehiculo else 'N/A'})" if instance.vehiculo else "Sin vehículo"
        total_info = f"Bs. {instance.total:.2f}"
        num_detalles = instance.detalles.count()
        
        # Registrar en bitácora
        descripcion = f"Presupuesto #{instance.id} creado para cliente '{cliente_info}', vehículo '{vehiculo_info}', total: {total_info}, con {num_detalles} item(s), estado: {instance.estado}"
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.CREAR,
            modulo=Bitacora.Modulo.PRESUPUESTO,
            descripcion=descripcion,
            request=self.request
        )

    def perform_update(self, serializer):
        """Actualizar presupuesto y registrar en bitácora"""
        # Guardar datos originales para comparación
        instance = self.get_object()
        estado_original = instance.estado
        cliente_original = f"{instance.cliente.nombre} {instance.cliente.apellido}" if instance.cliente else "Sin cliente"
        vehiculo_original = instance.vehiculo.numero_placa if instance.vehiculo else "Sin vehículo"
        total_original = instance.total
        con_impuestos_original = instance.con_impuestos
        
        # Ejecutar la actualización original
        instance = serializer.save()
        
        # Crear descripción detallada
        cambios = []
        if instance.estado != estado_original:
            cambios.append(f"estado: '{estado_original}' → '{instance.estado}'")
        
        cliente_nuevo = f"{instance.cliente.nombre} {instance.cliente.apellido}" if instance.cliente else "Sin cliente"
        if cliente_nuevo != cliente_original:
            cambios.append(f"cliente: '{cliente_original}' → '{cliente_nuevo}'")
        
        vehiculo_nuevo = instance.vehiculo.numero_placa if instance.vehiculo else "Sin vehículo"
        if vehiculo_nuevo != vehiculo_original:
            cambios.append(f"vehículo: '{vehiculo_original}' → '{vehiculo_nuevo}'")
        
        if instance.total != total_original:
            cambios.append(f"total: 'Bs. {total_original:.2f}' → 'Bs. {instance.total:.2f}'")
        
        if instance.con_impuestos != con_impuestos_original:
            cambios.append(f"con impuestos: '{con_impuestos_original}' → '{instance.con_impuestos}'")
        
        descripcion = f"Presupuesto #{instance.id} actualizado"
        if cambios:
            descripcion += f". Cambios: {', '.join(cambios)}"
        else:
            descripcion += ". Sin cambios detectados en campos principales"
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.EDITAR,
            modulo=Bitacora.Modulo.PRESUPUESTO,
            descripcion=descripcion,
            request=self.request
        )

    def perform_destroy(self, instance):
        """Eliminar presupuesto y registrar en bitácora"""
        # Guardar información antes de eliminar
        presupuesto_id = instance.id
        cliente_info = f"{instance.cliente.nombre} {instance.cliente.apellido}" if instance.cliente else "Sin cliente"
        vehiculo_info = instance.vehiculo.numero_placa if instance.vehiculo else "Sin vehículo"
        total_info = f"Bs. {instance.total:.2f}"
        estado_info = instance.estado
        num_detalles = instance.detalles.count()
        
        # Ejecutar la eliminación original
        instance.delete()
        
        # Registrar en bitácora
        descripcion = f"Presupuesto #{presupuesto_id} eliminado. Tenía cliente: '{cliente_info}', vehículo: '{vehiculo_info}', total: {total_info}, estado: {estado_info}, con {num_detalles} item(s)"
        
        registrar_bitacora(
            usuario=self.request.user,
            accion=Bitacora.Accion.ELIMINAR,
            modulo=Bitacora.Modulo.PRESUPUESTO,
            descripcion=descripcion,
            request=self.request
        )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            self.perform_create(serializer)
            return Response(PresupuestoSerializer(serializer.instance).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            self.perform_update(serializer)
            return Response(PresupuestoSerializer(serializer.instance).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'message': 'Presupuesto eliminado correctamente.'}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['get'])
    def detalles(self, request, pk=None):
        presup = self.get_object()
        detalles = presup.detalles.all()
        serializer = DetallePresupuestoSerializer(detalles, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """Exportar presupuesto individual a PDF - Versión con Canvas"""
        try:
            from reportlab.pdfgen import canvas as pdf_canvas
            from reportlab.lib.pagesizes import letter
            
            logger.info(f"=== Iniciando generación de PDF para presupuesto {pk} ===")
            presup = self.get_object()
            logger.info(f"Presupuesto obtenido: ID={presup.id}, Estado={presup.estado}")
            
            # Crear el PDF en memoria usando canvas directamente
            buffer = BytesIO()
            c = pdf_canvas.Canvas(buffer, pagesize=letter)
            width, height = letter
            
            logger.info(f"Canvas creado. Width={width}, Height={height}")
            
            # Título
            c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(width/2, height - 50, f"PRESUPUESTO #{presup.id}")
            
            # Información general
            y = height - 100
            c.setFont("Helvetica-Bold", 12)
            c.drawString(50, y, "Estado:")
            c.setFont("Helvetica", 12)
            estado_text = str(presup.estado).upper() if presup.estado else 'N/A'
            c.drawString(150, y, estado_text)
            logger.info(f"Estado dibujado: {estado_text}")
            
            y -= 20
            c.setFont("Helvetica-Bold", 12)
            c.drawString(50, y, "Fecha Inicio:")
            c.setFont("Helvetica", 12)
            c.drawString(150, y, presup.fecha_inicio.strftime('%d/%m/%Y') if presup.fecha_inicio else 'N/A')
            
            y -= 20
            c.setFont("Helvetica-Bold", 12)
            c.drawString(50, y, "Fecha Fin:")
            c.setFont("Helvetica", 12)
            c.drawString(150, y, presup.fecha_fin.strftime('%d/%m/%Y') if presup.fecha_fin else 'N/A')
            
            # Cliente
            y -= 40
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, y, "CLIENTE")
            
            y -= 25
            c.setFont("Helvetica", 12)
            cliente_nombre = f"{presup.cliente.nombre} {presup.cliente.apellido}" if presup.cliente else 'Sin cliente'
            c.drawString(50, y, cliente_nombre)
            
            # Vehículo
            if presup.vehiculo:
                y -= 40
                c.setFont("Helvetica-Bold", 14)
                c.drawString(50, y, "VEHICULO")
                
                y -= 25
                c.setFont("Helvetica", 12)
                c.drawString(50, y, f"Placa: {presup.vehiculo.numero_placa or 'N/A'}")
                
                y -= 20
                marca_nombre = presup.vehiculo.marca.nombre if presup.vehiculo.marca else 'N/A'
                c.drawString(50, y, f"Marca: {marca_nombre}")
                
                y -= 20
                modelo_nombre = presup.vehiculo.modelo.nombre if presup.vehiculo.modelo else 'N/A'
                c.drawString(50, y, f"Modelo: {modelo_nombre}")
            
            # Detalles
            y -= 40
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, y, "DETALLES DEL PRESUPUESTO")
            
            y -= 30
            # Headers de tabla
            c.setFont("Helvetica-Bold", 10)
            c.drawString(50, y, "Item")
            c.drawString(250, y, "Cant.")
            c.drawString(310, y, "Precio")
            c.drawString(390, y, "Desc.")
            c.drawString(460, y, "Subtotal")
            
            # Línea debajo de headers
            c.line(50, y-5, 545, y-5)
            
            y -= 20
            c.setFont("Helvetica", 9)
            
            for detalle in presup.detalles.all():
                if y < 100:  # Nueva página si no hay espacio
                    c.showPage()
                    y = height - 50
                    c.setFont("Helvetica", 9)
                
                item_nombre = detalle.item.nombre if detalle.item else 'N/A'
                if len(item_nombre) > 30:
                    item_nombre = item_nombre[:27] + "..."
                
                cantidad = str(detalle.cantidad) if detalle.cantidad else '0'
                precio = f"Bs. {float(detalle.precio_unitario):.2f}" if detalle.precio_unitario else "Bs. 0.00"
                descuento = f"{float(detalle.descuento_porcentaje):.0f}%" if detalle.descuento_porcentaje else "0%"
                subtotal = f"Bs. {float(detalle.total):.2f}" if detalle.total else "Bs. 0.00"
                
                c.drawString(50, y, item_nombre)
                c.drawString(250, y, cantidad)
                c.drawString(310, y, precio)
                c.drawString(390, y, descuento)
                c.drawString(460, y, subtotal)
                
                y -= 18
            
            # Totales
            y -= 20
            c.line(350, y, 545, y)
            
            y -= 20
            c.setFont("Helvetica", 11)
            
            subtotal_value = float(presup.subtotal) if presup.subtotal else 0.0
            descuentos_value = float(presup.total_descuentos) if presup.total_descuentos else 0.0
            total_value = float(presup.total) if presup.total else 0.0
            
            c.drawString(380, y, "Subtotal:")
            c.drawRightString(540, y, f"Bs. {subtotal_value:.2f}")
            
            y -= 20
            c.drawString(380, y, "Descuentos:")
            c.drawRightString(540, y, f"-Bs. {descuentos_value:.2f}")
            
            if presup.con_impuestos and presup.impuestos:
                impuestos_value = float(presup.impuestos)
                base_imponible = total_value / (1 + impuestos_value / 100)
                monto_impuesto = total_value - base_imponible
                
                y -= 20
                c.drawString(380, y, f"IVA ({impuestos_value:.0f}%):")
                c.drawRightString(540, y, f"Bs. {monto_impuesto:.2f}")
            
            y -= 20
            c.setFont("Helvetica-Bold", 12)
            c.drawString(380, y, "TOTAL:")
            c.drawRightString(540, y, f"Bs. {total_value:.2f}")
            
            # Diagnóstico
            if presup.diagnostico:
                y -= 40
                if y < 150:  # Nueva página si no hay espacio
                    c.showPage()
                    y = height - 50
                
                c.setFont("Helvetica-Bold", 12)
                c.drawString(50, y, "DIAGNOSTICO:")
                
                y -= 20
                c.setFont("Helvetica", 10)
                
                # Dividir diagnóstico en líneas
                diagnostico_text = str(presup.diagnostico)
                max_chars = 90
                lines = []
                
                while len(diagnostico_text) > max_chars:
                    split_pos = diagnostico_text[:max_chars].rfind(' ')
                    if split_pos == -1:
                        split_pos = max_chars
                    lines.append(diagnostico_text[:split_pos])
                    diagnostico_text = diagnostico_text[split_pos:].strip()
                
                if diagnostico_text:
                    lines.append(diagnostico_text)
                
                for line in lines:
                    if y < 50:
                        c.showPage()
                        y = height - 50
                        c.setFont("Helvetica", 10)
                    c.drawString(50, y, line)
                    y -= 15
            
            # Finalizar PDF
            logger.info("Llamando a c.save()...")
            c.save()
            logger.info("c.save() completado exitosamente")
            
            # Preparar respuesta
            buffer.seek(0)
            pdf_data = buffer.read()
            buffer_size = len(pdf_data)
            buffer.close()
            
            logger.info(f"Tamaño del PDF generado: {buffer_size} bytes")
            
            if buffer_size == 0:
                logger.error("ERROR: El PDF está vacío!")
                return Response(
                    {'error': 'El PDF generado está vacío'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            if buffer_size < 100:
                logger.error(f"ERROR: El PDF es demasiado pequeño ({buffer_size} bytes), probablemente corrupto")
                return Response(
                    {'error': f'El PDF generado parece corrupto (solo {buffer_size} bytes)'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            logger.info("Creando HttpResponse...")
            response = HttpResponse(pdf_data, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="presupuesto_{presup.id}.pdf"'
            response['Content-Length'] = buffer_size
            logger.info(f"HttpResponse creado. Headers: {dict(response.items())}")
            
            # Registrar en bitácora
            registrar_bitacora(
                usuario=request.user,
                accion=Bitacora.Accion.CREAR,
                modulo=Bitacora.Modulo.PRESUPUESTO,
                descripcion=f"Reporte PDF generado para Presupuesto #{presup.id}",
                request=request
            )
            
            logger.info("PDF enviado correctamente")
            return response
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            logger.error(f"ERROR COMPLETO al generar PDF:\n{error_detail}")
            return Response(
                {'error': f'Error al generar PDF: {str(e)}', 'detail': error_detail}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """Exportar presupuesto individual a Excel"""
        try:
            presup = self.get_object()
            
            # Contenedor para los elementos
            elements = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=12,
                spaceBefore=12
            )
            
            # Título
            elements.append(Paragraph(f"PRESUPUESTO #{presup.id}", title_style))
            elements.append(Spacer(1, 12))
            
            # Información general
            info_data = [
                ['Estado:', presup.estado.upper() if presup.estado else 'N/A'],
                ['Fecha Inicio:', presup.fecha_inicio.strftime('%d/%m/%Y') if presup.fecha_inicio else 'N/A'],
                ['Fecha Fin:', presup.fecha_fin.strftime('%d/%m/%Y') if presup.fecha_fin else 'N/A'],
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 20))
            
            # Cliente
            elements.append(Paragraph("INFORMACIÓN DEL CLIENTE", heading_style))
            cliente_nombre = f"{presup.cliente.nombre} {presup.cliente.apellido}" if presup.cliente else 'Sin cliente'
            cliente_data = [
                ['Cliente:', cliente_nombre]
            ]
            
            cliente_table = Table(cliente_data, colWidths=[2*inch, 4*inch])
            cliente_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            
            elements.append(cliente_table)
            elements.append(Spacer(1, 20))
            
            # Vehículo
            if presup.vehiculo:
                elements.append(Paragraph("INFORMACIÓN DEL VEHÍCULO", heading_style))
                vehiculo_data = [
                    ['Placa:', presup.vehiculo.numero_placa or 'N/A'],
                    ['Marca:', presup.vehiculo.marca.nombre if presup.vehiculo.marca else 'N/A'],
                    ['Modelo:', presup.vehiculo.modelo.nombre if presup.vehiculo.modelo else 'N/A'],
                ]
                
                vehiculo_table = Table(vehiculo_data, colWidths=[2*inch, 4*inch])
                vehiculo_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey)
                ]))
                
                elements.append(vehiculo_table)
                elements.append(Spacer(1, 20))
            
            # Detalles
            elements.append(Paragraph("DETALLES DEL PRESUPUESTO", heading_style))
            
            detalles_data = [['Item', 'Cant.', 'Precio Unit.', 'Desc.', 'Subtotal']]
            for detalle in presup.detalles.all():
                item_nombre = detalle.item.nombre if detalle.item else 'N/A'
                cantidad = str(detalle.cantidad) if detalle.cantidad else '0'
                precio = f"Bs. {float(detalle.precio_unitario):.2f}" if detalle.precio_unitario else "Bs. 0.00"
                descuento = f"{float(detalle.descuento_porcentaje):.0f}%" if detalle.descuento_porcentaje else "0%"
                subtotal = f"Bs. {float(detalle.total):.2f}" if detalle.total else "Bs. 0.00"
                detalles_data.append([item_nombre, cantidad, precio, descuento, subtotal])
            
            detalles_table = Table(detalles_data, colWidths=[2.5*inch, 0.8*inch, 1.2*inch, 0.8*inch, 1.2*inch])
            detalles_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            elements.append(detalles_table)
            elements.append(Spacer(1, 20))
            
            # Totales
            subtotal_value = float(presup.subtotal) if presup.subtotal else 0.0
            descuentos_value = float(presup.total_descuentos) if presup.total_descuentos else 0.0
            total_value = float(presup.total) if presup.total else 0.0
            
            totales_data = [
                ['Subtotal:', f"Bs. {subtotal_value:.2f}"],
                ['Descuentos:', f"-Bs. {descuentos_value:.2f}"],
            ]
            
            if presup.con_impuestos and presup.impuestos:
                impuestos_value = float(presup.impuestos)
                base_imponible = total_value / (1 + impuestos_value / 100)
                monto_impuesto = total_value - base_imponible
                totales_data.append([f'IVA ({impuestos_value:.0f}%):', f"Bs. {monto_impuesto:.2f}"])
            
            totales_data.append(['TOTAL:', f"Bs. {total_value:.2f}"])
            
            totales_table = Table(totales_data, colWidths=[4.5*inch, 2*inch])
            totales_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -2), 'Helvetica'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e5e7eb')),
            ]))
            
            elements.append(totales_table)
            
            # Diagnóstico
            if presup.diagnostico:
                elements.append(Spacer(1, 20))
                elements.append(Paragraph("DIAGNÓSTICO", heading_style))
                # Escapar caracteres especiales correctamente (& primero)
                diagnostico_text = str(presup.diagnostico)
                diagnostico_text = diagnostico_text.replace('&', '&amp;')
                diagnostico_text = diagnostico_text.replace('<', '&lt;')
                diagnostico_text = diagnostico_text.replace('>', '&gt;')
                # Usar Preformatted en lugar de Paragraph para evitar problemas
                elements.append(Preformatted(diagnostico_text, styles['Normal'], maxLineLength=80))
            
            # Construir PDF
            try:
                logger.info(f"Construyendo PDF con {len(elements)} elementos")
                doc.build(elements)
                logger.info("PDF construido exitosamente")
            except Exception as build_error:
                logger.error(f"Error al construir PDF: {str(build_error)}")
                return Response(
                    {'error': f'Error al construir PDF: {str(build_error)}'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # Preparar respuesta
            buffer.seek(0)
            pdf_data = buffer.read()
            buffer.close()
            
            logger.info(f"PDF generado, tamaño: {len(pdf_data)} bytes")
            
            if len(pdf_data) == 0:
                logger.error("PDF vacío generado")
                return Response(
                    {'error': 'El PDF generado está vacío'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            response = HttpResponse(pdf_data, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="presupuesto_{presup.id}.pdf"'
            
            # Registrar en bitácora
            registrar_bitacora(
                usuario=request.user,
                accion=Bitacora.Accion.CREAR,
                modulo=Bitacora.Modulo.PRESUPUESTO,
                descripcion=f"Reporte PDF generado para Presupuesto #{presup.id}",
                request=request
            )
            
            return response
            
        except Exception as e:
            logger.error(f"Error general al generar PDF: {str(e)}", exc_info=True)
            return Response(
                {'error': f'Error al generar PDF: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """Exportar presupuesto individual a Excel"""
        try:
            presup = self.get_object()
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Presupuesto {presup.id}"
            
            # Estilos
            header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=16, color="1e40af")
            bold_font = Font(bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:E1')
            ws['A1'] = f"PRESUPUESTO #{presup.id}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Información general
            row = 3
            ws[f'A{row}'] = "Estado:"
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = presup.estado.upper() if presup.estado else 'N/A'
            
            row += 1
            ws[f'A{row}'] = "Fecha Inicio:"
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = presup.fecha_inicio.strftime('%d/%m/%Y') if presup.fecha_inicio else 'N/A'
            
            row += 1
            ws[f'A{row}'] = "Fecha Fin:"
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = presup.fecha_fin.strftime('%d/%m/%Y') if presup.fecha_fin else 'N/A'
            
            # Cliente
            row += 2
            ws[f'A{row}'] = "CLIENTE"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            
            row += 1
            cliente_nombre = f"{presup.cliente.nombre} {presup.cliente.apellido}" if presup.cliente else 'Sin cliente'
            ws[f'A{row}'] = cliente_nombre
            
            # Vehículo
            if presup.vehiculo:
                row += 2
                ws[f'A{row}'] = "VEHÍCULO"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                
                row += 1
                ws[f'A{row}'] = "Placa:"
                ws[f'A{row}'].font = bold_font
                ws[f'B{row}'] = presup.vehiculo.numero_placa or 'N/A'
                
                row += 1
                ws[f'A{row}'] = "Marca:"
                ws[f'A{row}'].font = bold_font
                ws[f'B{row}'] = presup.vehiculo.marca.nombre if presup.vehiculo.marca else 'N/A'
                
                row += 1
                ws[f'A{row}'] = "Modelo:"
                ws[f'A{row}'].font = bold_font
                ws[f'B{row}'] = presup.vehiculo.modelo.nombre if presup.vehiculo.modelo else 'N/A'
            
            # Detalles
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = "DETALLES DEL PRESUPUESTO"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            headers = ['Item', 'Cantidad', 'Precio Unit.', 'Desc. (%)', 'Subtotal']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Datos de detalles
            for detalle in presup.detalles.all():
                row += 1
                ws.cell(row=row, column=1).value = detalle.item.nombre if detalle.item else 'N/A'
                ws.cell(row=row, column=2).value = float(detalle.cantidad) if detalle.cantidad else 0.0
                ws.cell(row=row, column=3).value = float(detalle.precio_unitario) if detalle.precio_unitario else 0.0
                ws.cell(row=row, column=4).value = float(detalle.descuento_porcentaje) if detalle.descuento_porcentaje else 0.0
                ws.cell(row=row, column=5).value = float(detalle.total) if detalle.total else 0.0
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                    if col >= 2:
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
            
            # Totales
            subtotal_value = float(presup.subtotal) if presup.subtotal else 0.0
            descuentos_value = float(presup.total_descuentos) if presup.total_descuentos else 0.0
            total_value = float(presup.total) if presup.total else 0.0
            
            row += 2
            ws[f'D{row}'] = "Subtotal:"
            ws[f'D{row}'].font = bold_font
            ws[f'E{row}'] = subtotal_value
            ws[f'E{row}'].alignment = Alignment(horizontal='right')
            
            row += 1
            ws[f'D{row}'] = "Descuentos:"
            ws[f'D{row}'].font = bold_font
            ws[f'E{row}'] = -descuentos_value
            ws[f'E{row}'].alignment = Alignment(horizontal='right')
            
            if presup.con_impuestos and presup.impuestos:
                impuestos_value = float(presup.impuestos)
                base_imponible = total_value / (1 + impuestos_value / 100)
                monto_impuesto = total_value - base_imponible
                row += 1
                ws[f'D{row}'] = f"IVA ({impuestos_value:.0f}%):"
                ws[f'D{row}'].font = bold_font
                ws[f'E{row}'] = monto_impuesto
                ws[f'E{row}'].alignment = Alignment(horizontal='right')
            
            row += 1
            ws[f'D{row}'] = "TOTAL:"
            ws[f'D{row}'].font = Font(bold=True, size=12)
            ws[f'E{row}'] = total_value
            ws[f'E{row}'].font = Font(bold=True, size=12)
            ws[f'E{row}'].alignment = Alignment(horizontal='right')
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            # Guardar en memoria
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            excel_data = buffer.read()
            buffer.close()
            
            # Preparar respuesta
            response = HttpResponse(
                excel_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="presupuesto_{presup.id}.xlsx"'
            
            # Registrar en bitácora
            registrar_bitacora(
                usuario=request.user,
                accion=Bitacora.Accion.CREAR,
                modulo=Bitacora.Modulo.PRESUPUESTO,
                descripcion=f"Reporte Excel generado para Presupuesto #{presup.id}",
                request=request
            )
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar Excel: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

